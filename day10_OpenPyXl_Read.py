import openpyxl

wb = openpyxl.load_workbook("student_level_example.xlsx","r")

sheet = wb['Sheet1']
rows=sheet.max_row
print(rows)
cols=sheet.max_colum
print(cols)

for i in range(1,rows+1):
    for j in range(1,cols+1):
        print(sheet.cell(i,j).value,end=" ")
    print()
print(sheet.cell(1,1).value)
